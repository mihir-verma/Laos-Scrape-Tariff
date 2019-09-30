import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import MVermaResLaos as M

M.Hello()
M.yesnoloop('Have you created a folder \'Output\\Laos\\\'? (y/n) ')

service = webdriver.chrome.service.Service(r'C:/Program Files (x86)/SeleniumWrapper/chromedriver.exe')
service.start()
chrome_options = Options()
chrome_options.headless = True
chrome_options.add_argument('--lang=en-US')
chrome_options.add_argument('--start-fullscreen')
#chrome_options.add_argument('--log-level=3')
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

driver = webdriver.Remote(service.service_url, desired_capabilities = chrome_options.to_capabilities())
print('\nGetting Laos Standards Page')
driver.get('https://www.laotradeportal.gov.la/index.php?r=searchMeasures/standard')
print('Selecting Language as English')
M.to_eng(driver)

desc = driver.find_element_by_xpath('//*[@id="measures-grid"]/div[1]').text
total_num, pages = M.find_pages(desc)

frame = pd.DataFrame(columns=list(range(15)))
frame.loc[0] = 'ID','Name','Description','Comments','Validity From','Validity To','Reference','Technical Code','Enforced By/Agency','Created Date','Updated Date','Status','Measure Type','Legal Document/Regulation','Total HS Codes'

driver2 = webdriver.Remote(service.service_url, desired_capabilities = chrome_options.to_capabilities())

for i in range(1,pages+1): #(1,pages+1):

    try:
        id_list_element = driver.find_element_by_xpath('//*[@id="measures-grid"]/div[3]')
    except:
        id_list_element = driver.find_element_by_xpath('//*[@id="measures-grid"]/div[2]')

    page_nav = driver.find_element_by_xpath('//*[@id="yw0"]')
    bi = page_nav.find_element_by_partial_link_text(str(i))
    driver.execute_script('arguments[0].click();', bi)
    
    wait = WebDriverWait(driver,25)
    element = wait.until(EC.staleness_of(id_list_element))
    
    try:
        id_list_element = driver.find_element_by_xpath('//*[@id="measures-grid"]/div[3]')
    except:
        id_list_element = driver.find_element_by_xpath('//*[@id="measures-grid"]/div[2]')
         
    id_elements = id_list_element.find_elements_by_xpath('./*')

    for id_element in id_elements:
        id = id_element.get_attribute('innerHTML')
        
        print('\nOpening page for ID', id, 'for fetching Total HS Codes information')
        driver2.get(M.link_maker(id,1))
        print('Selecting English Option')
        M.to_eng(driver2)

        print('Fetching Information')    
        
        desc2 = driver2.find_element_by_xpath('//*[@id="commodity-description-list"]/div[1]').text
        if desc2 == "":
            pages = 1
            total_num2=[0]
        else:
            total_num2, pages = M.find_pages(desc2)
        
        f = lambda x: driver2.find_element_by_xpath('//*[@id="yw0"]/tbody/tr['+str(x)+']/td').text
                
        frame.loc[len(frame)+1] = f(1),f(2),f(3),f(4),f(5),f(6),f(7),f(8),f(9),f(10),f(11),f(12),f(13),f(14),total_num2[0]    
        
    print('\nSaving the Output File')
    final_df = pd.concat([M.blanks(1), pd.Series([str(total_num[0])+' total Sub-Ministries']), M.blanks(1),frame], ignore_index=True)
    final_df.to_excel('../Output/Laos/Laos_Standards.xlsx', sheet_name='Main', header=False,index=False)

driver.quit()

print('\nChecking File for HS Codes')
    
all_ids=[]

path=r'../Output/Laos/Laos_Standards.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl')    
workbook=load_workbook(path)
sheet = workbook['Main']

print('Loading ID information')
row_end = sheet.max_row
        
for r in range(5,row_end+1):
    id = sheet.cell(row=r, column=15)
    if id.value!=0:
        all_ids.append(sheet.cell(row=r, column=1).value)

workbook.close()
    
for i in range(len(all_ids)): #(0,len(all_ids)):
            
    print('\nGetting Webpage with ID', all_ids[i])
    driver = webdriver.Remote(service.service_url, desired_capabilities = chrome_options.to_capabilities())
    driver.get(M.link_maker(all_ids[i],1))
    print('Selecting language as English')
    M.to_eng(driver)
                    
    desc = driver.find_element_by_xpath('//*[@id="commodity-description-list"]/div[1]').text
    if desc == "":
        pages = 1
        total_num=[0]
    else:
        total_num, pages = M.find_pages(desc)
    
    hs_codes= pd.DataFrame(columns=range(2))
    hs_codes.loc[0] = 'HS Codes', 'Description'
    
    for j in range(1,pages+1): #(1,pages+1):
        
        driver.get(M.link_maker(all_ids[i],j))
        page_hs_codes = M.get_table(driver, '//*[@id="commodity-description-list"]/table')
        page_hs_codes.drop(0, inplace=True)
        hs_codes = pd.concat([hs_codes,page_hs_codes], ignore_index=True)
        print('Information Fetched')
    
    print('\nSaving the File with a new Sheet for id', all_ids[i])
    try:
        writer.book = load_workbook(path)
        try:
            writer.book.remove(writer.book['Standard_id_'+str(all_ids[i])])
            writer.save()
        except:
            pass
    except:
        pass

    hs_codes.to_excel(writer, header=False,index=False, sheet_name='Standard_id_'+str(all_ids[i]))
    writer.save()
            
    driver.quit()

    writer.close()

M.complete_msg()