import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import MVermaResLaos as M

M.Hello()
M.yesnoloop('Have you run the first phase for Laos (Main Page Navigation)? (y/n) ')

try:
    init = int(input('\nGive the index from where to start extraction\n(Leave blank for default): '))
except:
    init = 1

service = webdriver.chrome.service.Service(r'C:/Program Files (x86)/SeleniumWrapper/chromedriver.exe')
service.start()
chrome_options = Options()
chrome_options.headless = True
chrome_options.add_argument('--lang=en-US')
#chrome_options.add_argument('--log-level=3')
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

for i in range(init,15):
    
    print('\nOpening File of Measure_Type', i, 'for processing')
    all_ids=[]
    
    path=r'../Output/Laos/Lao_MeasureType_'+str(i)+'.xlsx'
    writer = pd.ExcelWriter(path, engine='openpyxl')
    
    workbook=load_workbook(path)
    main_sheet = workbook['Main']
    
    print('Loading ID information')
    row_end = main_sheet.max_row
    
    for r in range(7,row_end+1):
        id = main_sheet.cell(row=r, column=3)
        all_ids.append(id.value)
    
    workbook.close()
    
    for j in range(0,len(all_ids)): #(0,len(all_ids)):
        
        print('\nGetting Webpage with ID', all_ids[j], '(Ministry Type Index', i, '\b)')
        driver = webdriver.Remote(service.service_url, desired_capabilities = chrome_options.to_capabilities())
        driver.get(M.link_maker(all_ids[j],1))
        print('Selecting language as English')
        M.to_eng(driver)
                        
        desc = driver.find_element_by_xpath('//*[@id="commodity-description-list"]/div[1]').text
        if desc == "":
            pages = 1
            total_num=[0]
        else:
            total_num, pages = M.find_pages(desc)
        
        print('Fetching Information')
        view_measures = M.get_table(driver, '//*[@id="yw0"]')
        procedures = M.get_table(driver, '//*[@id="procedure-list"]/table')

        hs_codes= pd.DataFrame(columns=range(2))
        hs_codes.loc[0] = 'HS Codes', 'Description'
        
        for k in range(1,pages+1): #(1,pages+1):
            
            print('\nOpening page', k, 'for ID', all_ids[j], '(Ministry Type Index', i, '\b)')
            driver.get(M.link_maker(all_ids[j],k))
            page_hs_codes = M.get_table(driver, '//*[@id="commodity-description-list"]/table')
            page_hs_codes.drop(0, inplace=True)
            hs_codes = pd.concat([hs_codes,page_hs_codes], ignore_index=True)
            print('Information Fetched')
        
        print('\nSaving the File with a new Sheet for id', all_ids[j])
        try:
            writer.book = load_workbook(path)
            try:
                writer.book.remove(writer.book['SubMinistry_id_'+str(all_ids[j])])
                writer.save()
            except:
                pass
        except:
            pass
            
        sheetdf = pd.concat([M.timestamp(), M.blanks(1), pd.Series(['View Measures',]),M.blanks(1),view_measures,M.blanks(2), pd.Series(['Procedures']), M.blanks(1),procedures,M.blanks(2),pd.Series([str(total_num[0])+' total HS Codes']), pd.Series(['HS Codes']),M.blanks(1), hs_codes], ignore_index=True)
        sheetdf.to_excel(writer, header=False,index=False, sheet_name='SubMinistry_id_'+str(all_ids[j]))
        writer.save()
        
        print('\n'+M.timestamp().loc[0,1]+'\n')
        
        driver.quit()
    
        writer.close()

M.complete_msg()